import openpyxl
import os

pos = ['명사','동사']
tag = ['준말','높임말','유의어']

dictionary = openpyxl.load_workbook("한국어기초사전 관련어.xlsx")
sheet = dictionary.active

for p in pos:
    for t in tag:
        wb = openpyxl.Workbook()
        ws = wb.active
        row_index = 1

        for row in sheet.rows:
            if row[3].value == p and t in row[4].value:
                word = row[4].value.split(t +' ')
                if len(word) == 2:
                    ws.cell(row=row_index, column=1, value = row[0].value)
                    ws.cell(row=row_index, column=2, value = word[1].split('참고어')[0].split('반대말')[0].split('준말')[0].split('센말')[0].split('높임말')[0].split('유의어')[0])
                    row_index += 1

        wb.save("./" + p + "_" + t + ".xlsx")
        wb.close()
dictionary.close()
